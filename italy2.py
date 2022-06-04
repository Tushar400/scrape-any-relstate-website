from statistics import mode
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import undetected_chromedriver as uc
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import csv
import random
from openpyxl import Workbook
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
import os
import urllib
import math
import sys
import time
import requests
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import tkinter as tk
from tkinter import simpledialog
def chrome_webdriver():    
   

    all_data=[]
   

    for i in range(3):
        
        page_links=[]
        for indexer,i in enumerate(driver.find_elements(By.XPATH,'//table[@class="table tableRisultatiGratuita"]/tbody/tr')):
            page_links.append(driver.find_elements(By.XPATH,'//table[@class="table tableRisultatiGratuita"]/tbody/tr')[indexer].find_element(By.TAG_NAME,'a').get_attribute('href'))
            
        for indexer,i in enumerate(driver.find_elements(By.XPATH,'//table[@class="table tableRisultatiGratuita"]/tbody/tr')):
            current_data=driver.find_elements(By.XPATH,'//table[@class="table tableRisultatiGratuita"]/tbody/tr')[indexer].text.splitlines()
            if len(current_data)==7:
                current_data.remove('... Leggi tutto')
            if len(current_data)==5:
                current_data.insert(-2,'Nothing')
            
            parent = driver.window_handles[0]
            driver.execute_script(f'''window.open("{page_links[indexer]}","_blank");''')
            chld = driver.window_handles[1]
            driver.switch_to.window(chld)
            #driver.find_elements_by_xpath('//table[@class="table tableRisultatiGratuita"]/tbody/tr')[indexer].find_element_by_tag_name('img').click()
            
            temp_list=[]
            for j in driver.find_elements(By.TAG_NAME,'dd'):
                temp_list.append(j.text.replace('Mostra mappa','').strip()) 
            try:
                temp_list.remove('')
            except:
                pass
            if 'MOSTRA' in temp_list:
                current_data.insert(1,temp_list[1])
                current_data.append(temp_list[-3])
                current_data.append(temp_list[-2])
                current_data.append(temp_list[-1])
                current_data[5]=temp_list[4]
                email=(driver.find_element(By.XPATH,'//a[@class="linkRisultatiRicerca showModalPec"]').get_attribute('onclick')).split('(')[-1]
                email=email.replace(')','')
                email=email.replace("'",'').strip()
                current_data.append(email)
                print(current_data)
            elif 'Non presente' in temp_list:
                current_data.insert(1,temp_list[1])
                current_data.append(temp_list[-3])
                current_data.append(temp_list[-2])
                current_data.append(temp_list[-1])
                current_data[5]=temp_list[4]
                email='Non presente'
                current_data.append(email)
                print(current_data)
            else:
                current_data.insert(1,temp_list[1])
                current_data.append(temp_list[-3])
                current_data.append(temp_list[-2])
                current_data.append(temp_list[-1])
                current_data[5]=temp_list[3]
                email='Non presente'
                current_data.append(email)
                print(current_data)      
            print(len(current_data))
            if len(current_data)!=11:
                raise Exception('More items')
            all_data.append(current_data)
            driver.close()
            driver.switch_to.window(parent)
        print(len(all_data))
        
        driver.execute_script("scrollBy(0,-1000000000000000);")
        driver.find_element(By.XPATH,"//a[contains(text(), 'Successivo')]").click()
    print('Data extraction completed')
    time.sleep(2)
    bold_font = Font(bold=True)
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    workbook = Workbook()
    sheet = workbook.active
    #['Pompe Funebri Sarni', 'Vallata (AV) Corso Kennedy 70/A', 'Sede Legale', 'Avellino', 'Altre Forme', "Servizi di pompe funebri e attivita' connesse", 'Registrata', '96.03', '96.03', '-', 'sarnigerardo80@sicurezzapostale.it']
    sheet["A1"] = "Nome impresa"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Indirizzo Impresa"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Stato dell'ufficio"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Provincia"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Forma giuridica (generico)"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Attività"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Registrato/Non registrato"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border
    sheet["H1"] = "ATECO prevalente"
    sheet["H1"].font = bold_font
    sheet["H1"].alignment = center_aligned_text
    sheet["H1"].border = square_border
    sheet["I1"] = "ATECO primaria"
    sheet["I1"].font = bold_font
    sheet["I1"].alignment = center_aligned_text
    sheet["I1"].border = square_border
    sheet["J1"] = "ATECO secondaria"
    sheet["J1"].font = bold_font
    sheet["J1"].alignment = center_aligned_text
    sheet["J1"].border = square_border
    sheet["K1"] = "DOMICILIO DIGITALE / PEC*"
    sheet["K1"].font = bold_font
    sheet["K1"].alignment = center_aligned_text
    sheet["K1"].border = square_border

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    for p,q in enumerate(all_data):
        sheet[f"A{p+2}"]=q[0]           
        sheet[f"B{p+2}"]=q[1]
        sheet[f"C{p+2}"]=q[2]
        sheet[f"D{p+2}"]=q[3]
        sheet[f"E{p+2}"]=q[4]
        sheet[f"F{p+2}"]=q[5]
        sheet[f"G{p+2}"]=q[6]
        sheet[f"H{p+2}"]=q[7]
        sheet[f"I{p+2}"]=q[8]
        sheet[f"J{p+2}"]=q[9]
        sheet[f"K{p+2}"]=q[10]
    workbook.save(f"reg.xlsx")
if __name__ == "__main__":   
    ROOT = tk.Tk()

    ROOT.withdraw()
   
    
  

    options = webdriver.ChromeOptions() 
    options.add_argument("start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('--disable-blink-features=AutomationControlled')
    driver = webdriver.Chrome( chrome_options=options)
    options.add_argument("--headless")
    driver.get('https://www.registroimprese.it/ricerca-libera-e-acquisto?p_p_id=ricercaportlet_WAR_ricercaRIportlet&p_p_lifecycle=0&p_p_state=normal&_ricercaportlet_WAR_ricercaRIportlet_cur=1&_ricercaportlet_WAR_ricercaRIportlet_pageToken=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2NDk3NTQ3NDIsImNvdW50IjoyNTB9.b5MPQ9jnQpIp4mCONZH5FVgXmu8CMwo9JXegCevnJJY')
    
    time.sleep(60)
   
    
    chrome_webdriver()
    
    print('process completed')
